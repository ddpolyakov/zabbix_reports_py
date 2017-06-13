from openpyxl import *
from openpyxl.styles import *
from zabbix.api import ZabbixAPI
from datetime import datetime
import os


date_text=str(datetime.now().strftime('%Y%m%d%H%M%S'))
zuser="" # имя пользователя
zpass="" # пароль пользователя
zurl="http://zabbix_egrn"   # адрес сервера
zapi = ZabbixAPI(url=zurl, user=zuser, password=zpass)
groupIds=['86','87'] # id групп хостов
ofile="./report"+date_text+".xlsx"
wb = Workbook ()

def get_data(g):
    for host in zapi.host.get(groupids=g, output="hostid"):
        hname = zapi.host.get(output=("hostid","name"),hostids=host)[0]['name']
        hid = zapi.host.get(output=("hostid","name"),hostids=host)[0]['hostid']
        dict1[hname]={}
        for items in zapi.item.get(output=["lastvalue","name"], hostids=hid):
            itname = items['name']
            if itname == "Время работы":
                m,s = divmod(int(items['lastvalue']), 60)
                h,m = divmod(m, 60)
                d,h = divmod(h, 24)
                itvalue = "%dд, %02d:%02d:%02d" % (d, h, m,s)
                print(itvalue)
            else:
                itvalue=items['lastvalue']
            #print(hname,itname,itvalue)
            dict1[hname][itname]=itvalue

def add_sheet(d,groupname):
    if wb.active.title == "Sheet":
        wb.remove_sheet(wb.active)
    ws=wb.create_sheet(title=groupname)
    i=0
    ws.merge_cells('A1:E1')
    ws['A1'] = "Группа хостов:  "+groupname
    for host in sorted(d):
        #print(host)
        ws.cell(column=1,row=3+i, value=host)
        j=0
        for key in sorted(d[host]):
            value=key
            ws.cell(column=2+j, row=2, value=value)
            ws.cell(column=2+j, row=2).font = Font(bold=True)
            ws.cell(column=2+j, row=3+i, value=d[host][key])
            j = j + 1
        i = i + 1
    data = ws.get_cell_collection()
    dict2 = dict()
    for i in data:
        key = i.column
        subkey = i.row
        value = i.coordinate
        if key in dict2:
            dict2[key][subkey] = value
        else:
            dict2[key] = {}
            dict2[key][subkey] = value
    #print(dict2)
    for col in dict2:
        data1 = []
        for cell in dict2[col]:
            data1.append(dict2[col][cell])
        column_widths = []
        for cell in data1:
            #print(cell)
            cellvalue = ws.cell(cell).value
            ws.cell(cell).alignment=Alignment(horizontal='center')
            column_widths.append(len(cellvalue))
            #print(max(column_widths))
            ws.column_dimensions[col].width = max(column_widths) + 5
    wb.save(ofile)

def init():
    try:
        os.remove(ofile)
    except OSError:
        pass
    for group in groupIds:
        global dict1
        dict1=dict()
        #print(group)
        groupname=zapi.hostgroup.get(output=("groupid","name"),groupids=group)[0]['name']
        get_data(group)
        add_sheet(dict1,groupname)
init()
